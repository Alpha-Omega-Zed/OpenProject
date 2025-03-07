import openpyxl
from bs4 import BeautifulSoup
from io import BytesIO
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

html_table = sys.stdin.read()

soup = BeautifulSoup(html_table, 'html.parser')
table = soup.find('table', {'class': 'report'})
header_rows = table.find_all('thead')
for header in header_rows:
    header.extract()
footer_rows = table.find_all('tfoot')
for footer in footer_rows:
    footer.extract()


cell_spans = []
cell_classes = []

rows=[row for header in header_rows for row in header.find_all('tr')] + table.find_all('tr') + [row for footer in footer_rows for row in footer.find_all('tr')]
for row in rows:
    row_data = []
    class_data = []
    for cell in row.find_all(['th', 'td']):
        colspan = int(cell.get('colspan', 1))
        rowspan = int(cell.get('rowspan', 1))

        cell_class = " ".join(cell.get('class', []))
        cell_text = cell.get_text(separator=" ", strip=True)  if cell_class!="normal inner left -break-word" else cell.get_text(separator="--_--", strip=True).split("--_--")[0]

        row_data.append((cell_text, colspan, rowspan))
        class_data.append((cell_class, colspan, rowspan))

    cell_spans.append(row_data)
    cell_classes.append(class_data)


class_colors = {
    'inner': 'F0F1F2',
    'top left -breakword': 'E9EDF0',
    'normal inner left -break-word 0': 'F0F1F2',
    'normal inner left -break-word 1': 'F6F8FA', 
    'top empty': 'FFFFFF', 
    'normal right': 'FFFFFF',
    'normal inner right 0': 'F0F1F2',
    'normal inner right 1': 'F6F8FA',
    'top right': 'E9EDF0',
    'bottom empty': 'FFFFFF',
    'bottom top empty': 'FFFFFF', 
    'normal empty': 'FFFFFF',
    'top': 'E9EDF0', 
    'top result': 'E9EDF0',
    'headerFooter':'E9EDF0'
}

# Assign white color to any remaining class
for row in cell_classes:
    for cell_class, _, _ in row:
        if cell_class and cell_class not in class_colors:
            class_colors[cell_class] = "FFFFFF"

output = BytesIO()

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Report'

merged_cells = {}
normalInnerLeftCount=0
normalInnerRightCount=0
for row_idx, (row_data, class_data) in enumerate(zip(cell_spans, cell_classes), start=1):
    col_idx = 1

    for (text, colspan, rowspan), (cell_class, _, _) in zip(row_data, class_data):
        while (row_idx, col_idx) in merged_cells:
            col_idx += 1

        worksheet.cell(row=row_idx, column=col_idx, value=text)
        # worksheet.cell(row=row_idx, column=col_idx).font = Font(name='Calibri',bold=True,size=11)

        if(cell_class=='normal inner left -break-word'):
            fill = PatternFill(start_color=class_colors[f'normal inner left -break-word {normalInnerLeftCount%2}'],
                               end_color=class_colors[f'normal inner left -break-word {normalInnerLeftCount%2}'],
                               fill_type="solid")
            normalInnerLeftCount+=1
        elif(cell_class=='normal inner right'):
            fill = PatternFill(start_color=class_colors[f'normal inner right {normalInnerRightCount%2}'],
                               end_color=class_colors[f'normal inner right {normalInnerRightCount%2}'],
                               fill_type="solid")
            normalInnerRightCount+=1
        elif cell_class and cell_class in class_colors:
            fill = PatternFill(start_color=class_colors[cell_class],
                               end_color=class_colors[cell_class],
                               fill_type="solid")
            # if class_colors[cell_class]=='FFFFFF':
            #     worksheet.cell(row=row_idx, column=col_idx).font = Font(name='Calibri',bold=False,size=11)
        else :
            fill = PatternFill(start_color=class_colors['headerFooter'],
                               end_color=class_colors['headerFooter'],
                               fill_type="solid")
            
        worksheet.cell(row=row_idx, column=col_idx).fill = fill


        if colspan > 1 or rowspan > 1:
            end_row = row_idx + rowspan - 1
            end_col = col_idx + colspan - 1
            worksheet.merge_cells(start_row=row_idx, start_column=col_idx,
                                  end_row=end_row, end_column=end_col)

            for r in range(row_idx, end_row + 1):
                for c in range(col_idx, end_col + 1):
                    merged_cells[(r, c)] = True

        col_idx += colspan

for col_idx in range(1, worksheet.max_column + 1):
    max_length = 0
    for row_idx in range(1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

for row_idx in range(1, worksheet.max_row + 1):
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.border = Border(
            left=Side(border_style="thin", color="D3D3D3"),
            right=Side(border_style="thin", color="D3D3D3"),
            top=Side(border_style="thin", color="D3D3D3"),
            bottom=Side(border_style="thin", color="D3D3D3")
        )

for row_idx in range(1, worksheet.max_row + 1):
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center")


workbook.save(output)
output.seek(0)

sys.stdout.buffer.write(output.getvalue())
